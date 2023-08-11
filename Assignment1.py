# Libraries that use in the program.
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook


# Initialize Chrome WebDriver
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--start-maximized")  # Maximize the browser window
driver = webdriver.Chrome(options=chrome_options)

# Read keywords from the Excel file
excel_file = "C:/Users/User/Desktop/Assgnment1Python/your_file.xlsx"
wb = load_workbook(excel_file)
sheet = wb.active

# Loop through keywords and perform searches
for row_num, row in enumerate(sheet.iter_rows(min_row=3, values_only=True, min_col=3, max_col=3), start=3):
    keyword = row[0]

    # Check if keyword is not None before stripping
    if keyword is not None:
        keyword = keyword.strip()

        # Open Google and find the search bar
        driver.get("https://www.google.com")
        search_box = driver.find_element("name", "q")

        # Search the keyword
        search_box.clear()
        search_box.send_keys(keyword)

        # Wait for auto-suggestions to appear
        time.sleep(2)
        suggestions = driver.find_elements(By.CSS_SELECTOR, ".erkvQe li")

        # Capture the suggested values
        suggested_values = [suggestion.find_element(By.XPATH, ".//div").text for suggestion in suggestions]

        # Get the longest and shortest suggested values
        longest_suggested = max(suggested_values, key=len)
        shortest_suggested = min(suggested_values, key=len)

        # Update results in the existing sheet
        sheet.cell(row=row_num, column=4, value=longest_suggested)  # Update Longest Suggested Value
        sheet.cell(row=row_num, column=5, value=shortest_suggested)  # Update Shortest Suggested Value

# Save the updated results to the Excel file
wb.save(excel_file)

# Close the browser window
driver.quit()
